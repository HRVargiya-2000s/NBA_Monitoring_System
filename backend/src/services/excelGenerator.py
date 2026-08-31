import sys
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter

def col_letter_of(col_idx):
    return get_column_letter(col_idx)

def main():
    if len(sys.argv) < 4:
        print("Usage: python excelGenerator.py <data_json_path> <template_path> <output_path>")
        sys.exit(1)

    data_json_path = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]

    with open(data_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(template_path)

    context = data.get('context', {})
    students = data.get('students', [])
    strength_mappings = data.get('strength_mappings', [])
    papers = data.get('papers', {})
    student_marks = data.get('student_marks', {})

    subject_name = context.get('subject_name', '')
    subject_code = context.get('subject_code', '')
    branch_name = context.get('branch_name', '')
    branch_code = context.get('branch_code', '')
    academic_year = context.get('accadmic_year', '')
    session = context.get('session', '')
    sem_number = context.get('sem_number', '')
    faculty_name = context.get('faculty_name', '')

    num_students = len(students)

    # 1. Update 'CO-PO Mapping' sheet
    if 'CO-PO Mapping' in wb.sheetnames:
        sheet = wb['CO-PO Mapping']
        sheet['A1'].value = subject_name
        sheet['A2'].value = subject_code
        sheet['B2'].value = f"{branch_code} - {branch_name}"

        # Write max marks for papers
        mid_sem_paper = papers.get('mid_sem', {})
        internal_paper = papers.get('internal', {})
        external_paper = papers.get('external', {})
        viva_paper = papers.get('viva', {})

        sheet['F17'].value = mid_sem_paper.get('max_marks') if mid_sem_paper.get('max_marks') is not None else ""
        sheet['H17'].value = internal_paper.get('max_marks') if internal_paper.get('max_marks') is not None else ""
        sheet['E17'].value = external_paper.get('max_marks') if external_paper.get('max_marks') is not None else ""
        sheet['G17'].value = viva_paper.get('max_marks') if viva_paper.get('max_marks') is not None else ""

        # Map Outcomes dynamically
        outcomes_list = []
        for i in range(1, 13):
            outcomes_list.append(('PO', i))
        for i in range(1, 5):
            outcomes_list.append(('PSO', i))

        # Write Outcome Headers (columns 3 to 18)
        for col_idx, (out_type, out_code) in enumerate(outcomes_list, 3):
            sheet.cell(row=33, column=col_idx).value = f"{out_type}{out_code}"

        # Build outcome coordinates lookup
        outcome_to_col = {}
        for col_idx, (out_type, out_code) in enumerate(outcomes_list, 3):
            outcome_to_col[f"{out_type}|{out_code}"] = col_idx

        # Clear strength mapping cells
        for r in range(34, 40):
            for c in range(3, 19):
                sheet.cell(row=r, column=c).value = None

        # Write Strength Mappings
        for mapping in strength_mappings:
            co_num = int(mapping.get('co_number', 0))
            out_type = mapping.get('outcome_type', '').upper()
            out_code = int(mapping.get('outcome_code', 0))
            strength = mapping.get('strength')

            key = f"{out_type}|{out_code}"
            col_idx = outcome_to_col.get(key)
            if col_idx and 1 <= co_num <= 6:
                row_idx = 33 + co_num
                # Write empty if strength is 0
                sheet.cell(row=row_idx, column=col_idx).value = strength if (strength and strength > 0) else None

        # Write formulas for Average
        for col_idx in range(3, 19):
            col_letter = col_letter_of(col_idx)
            sheet.cell(row=40, column=col_idx).value = f"=IFERROR(AVERAGE({col_letter}34:{col_letter}39),\"\")"

        # Write formulas for Attained
        for col_idx in range(3, 19):
            col_letter = col_letter_of(col_idx)
            sheet.cell(row=41, column=col_idx).value = f"=IFERROR(Attainment!{col_letter}27,\"\")"

        # Justifications
        # Columns D (4) and K (11) for COs. Clear first
        for col in [4, 11]:
            for r in range(46, 57):
                sheet.cell(row=r, column=col).value = None
            for r in range(58, 69):
                sheet.cell(row=r, column=col).value = None
            for r in range(70, 81):
                sheet.cell(row=r, column=col).value = None

        # Write Mapped Justifications (only PO1-PO11 supported in the 11 rows of template)
        for mapping in strength_mappings:
            co_num = int(mapping.get('co_number', 0))
            out_type = mapping.get('outcome_type', '').upper()
            out_code = int(mapping.get('outcome_code', 0))
            justification = mapping.get('justification', '')

            if out_type == 'PO' and 1 <= out_code <= 11 and justification:
                # Map to correct cell
                if co_num == 1:
                    sheet.cell(row=45 + out_code, column=4).value = justification
                elif co_num == 2:
                    sheet.cell(row=45 + out_code, column=11).value = justification
                elif co_num == 3:
                    sheet.cell(row=57 + out_code, column=4).value = justification
                elif co_num == 4:
                    sheet.cell(row=57 + out_code, column=11).value = justification
                elif co_num == 5:
                    sheet.cell(row=69 + out_code, column=4).value = justification
                elif co_num == 6:
                    sheet.cell(row=69 + out_code, column=11).value = justification

    # 2. Update 'Attainment' sheet
    if 'Attainment' in wb.sheetnames:
        sheet = wb['Attainment']
        sheet['A1'].value = "='CO-PO Mapping'!A1:N1"
        sheet['A2'].value = "='CO-PO Mapping'!A2"
        sheet['B2'].value = "='CO-PO Mapping'!B2:N2"

        # Write Outcome Headers (columns 3 to 18)
        outcomes_list = []
        for i in range(1, 13):
            outcomes_list.append(('PO', i))
        for i in range(1, 5):
            outcomes_list.append(('PSO', i))

        for col_idx, (out_type, out_code) in enumerate(outcomes_list, 3):
            sheet.cell(row=20, column=col_idx).value = f"{out_type}{out_code}"

        # Write matrix cells
        for co_num in range(1, 7):
            row_idx = 20 + co_num
            for c_idx in range(3, 19):
                c_letter = col_letter_of(c_idx)
                # Formula: =IF('CO-PO Mapping'!C$34="","",$H$9*'CO-PO Mapping'!C$34/3)
                sheet.cell(row=row_idx, column=c_idx).value = f"=IF('CO-PO Mapping'!{c_letter}${33 + co_num}=\"\",\"\",$H${8 + co_num}*'CO-PO Mapping'!{c_letter}${33 + co_num}/3)"

        # Write formulas for Average
        for col_idx in range(3, 19):
            col_letter = col_letter_of(col_idx)
            sheet.cell(row=27, column=col_idx).value = f"=IFERROR(AVERAGE({col_letter}21:{col_letter}26),\"\")"

    # 3. Update 'CO-PO-PSO' sheet and 'CO-PO Mapping' sheet with actual course outcomes descriptions and clean formatting
    from openpyxl.styles import Alignment
    wrap_align = Alignment(wrap_text=True, vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')

    course_outcomes = data.get('course_outcomes', [])
    mapped_cos = {int(co.get('co_number', 0)): co.get('co_description', '') for co in course_outcomes}

    if 'CO-PO-PSO' in wb.sheetnames:
        sheet = wb['CO-PO-PSO']
        for co_num in range(1, 7):
            co_desc = mapped_cos.get(co_num, "")
            c1 = sheet.cell(row=co_num, column=1)
            c2 = sheet.cell(row=co_num, column=2)

            c1.value = f"=CONCATENATE(\"CO \",'CO-PO Mapping'!$A$2,\".{co_num}\")" if co_desc else None
            c2.value = co_desc if co_desc else None

            c1.alignment = center_align
            c2.alignment = wrap_align
            if co_desc:
                sheet.row_dimensions[co_num].height = 45

        # Format PO and PSO rows cleanly
        for r in range(11, 32):
            cell = sheet.cell(row=r, column=2)
            if cell.value:
                cell.alignment = wrap_align

        if branch_name:
            sheet['A35'].value = f"PSO of {branch_name} Department"

        for r in range(36, 39):
            sheet.cell(row=r, column=1).alignment = center_align
            cell = sheet.cell(row=r, column=2)
            if cell.value:
                cell.alignment = wrap_align

    if 'CO-PO Mapping' in wb.sheetnames:
        sheet = wb['CO-PO Mapping']
        for co_num in range(1, 7):
            co_desc = mapped_cos.get(co_num, "")
            cell = sheet.cell(row=6 + co_num, column=3)
            cell.value = co_desc if co_desc else None
            cell.alignment = wrap_align
            if co_desc:
                sheet.row_dimensions[6 + co_num].height = 45

    # 4. Update 'CO-PO mapping summary' sheet
    if 'CO-PO mapping summary' in wb.sheetnames:
        sheet = wb['CO-PO mapping summary']
        sheet['C2'].value = f"{branch_name} Sem - {sem_number} ({academic_year})"
        sheet['C3'].value = f"Subject: {subject_code} - {subject_name}"

    # 5. Populate student roster in 'Students Details'
    if 'Students Details' in wb.sheetnames:
        sheet = wb['Students Details']
        # Clear student details range to prevent leftover data
        for r in range(7, 207):
            sheet.cell(row=r, column=1).value = None
            sheet.cell(row=r, column=2).value = None
            sheet.cell(row=r, column=3).value = None

        for idx, student in enumerate(students):
            row_idx = 7 + idx
            sheet.cell(row=row_idx, column=1).value = idx + 1
            
            enroll = student.get('enrollment_no', '')
            try:
                enroll = int(enroll)
            except ValueError:
                pass
            sheet.cell(row=row_idx, column=2).value = enroll
            sheet.cell(row=row_idx, column=3).value = student.get('name', '')

    # 6. Populate Marks and Formulas in M, I, E, V and Components
    # For Mid Sem
    mid_sem_paper = papers.get('mid_sem', {})
    mid_co_limits = mid_sem_paper.get('co_limits', {})
    mid_max_marks = mid_sem_paper.get('max_marks', 30)

    if 'M Component' in wb.sheetnames:
        sheet = wb['M Component']
        sheet['C4'].value = "='CO-PO Mapping'!B2"
        sheet['F4'].value = "='CO-PO Mapping'!A2"
        sheet['D5'].value = f"(Out of {mid_max_marks})"
        # Clear marks
        for r in range(8, 208):
            for c in range(4, 10):
                sheet.cell(row=r, column=c).value = None
        # Write student marks
        for idx, student in enumerate(students):
            row_idx = 8 + idx
            enroll = student.get('enrollment_no', '')
            s_marks = student_marks.get(enroll, {}).get('mid_sem', {})
            for co in range(1, 7):
                sheet.cell(row=row_idx, column=3 + co).value = s_marks.get(str(co))

    if 'M' in wb.sheetnames:
        sheet = wb['M']
        # Write max marks per CO
        for co in range(1, 7):
            sheet.cell(row=6, column=3 + co).value = mid_co_limits.get(str(co))
        sheet.cell(row=6, column=10).value = mid_max_marks

        # Set formulas for student rows
        for idx in range(num_students):
            r = 7 + idx
            sheet.cell(row=r, column=1).value = f"='Students Details'!A{r}"
            sheet.cell(row=r, column=2).value = f"=IF('Students Details'!B{r}=\"\",\"\",'Students Details'!B{r})"
            sheet.cell(row=r, column=3).value = f"=IF('Students Details'!C{r}=\"\",\"\",'Students Details'!C{r})"
            for co in range(1, 7):
                # 'M Component' K is column 11, matching 10 + co
                sheet.cell(row=r, column=3 + co).value = f"=IF('M Component'!{col_letter_of(10+co)}{8+idx}=0,\"\",'M Component'!{col_letter_of(10+co)}{8+idx})"
            sheet.cell(row=r, column=10).value = f"=SUM(D{r}:I{r})"

        # Clear remaining student rows formulas
        for r in range(7 + num_students, 207):
            for c in range(1, 11):
                sheet.cell(row=r, column=c).value = None

    # For Internal
    internal_paper = papers.get('internal', {})
    int_co_limits = internal_paper.get('co_limits', {})
    int_max_marks = internal_paper.get('max_marks', 20)

    if 'I Component' in wb.sheetnames:
        sheet = wb['I Component']
        sheet['C4'].value = "='CO-PO Mapping'!B2"
        sheet['F4'].value = "='CO-PO Mapping'!A2"
        sheet['D5'].value = f"(Out of {int_max_marks})"
        # Clear marks
        for r in range(8, 208):
            for c in range(4, 10):
                sheet.cell(row=r, column=c).value = None
        # Write student marks
        for idx, student in enumerate(students):
            row_idx = 8 + idx
            enroll = student.get('enrollment_no', '')
            s_marks = student_marks.get(enroll, {}).get('internal', {})
            for co in range(1, 7):
                sheet.cell(row=row_idx, column=3 + co).value = s_marks.get(str(co))

    if 'I' in wb.sheetnames:
        sheet = wb['I']
        # Write max marks per CO
        for co in range(1, 7):
            sheet.cell(row=6, column=3 + co).value = int_co_limits.get(str(co))
        sheet.cell(row=6, column=10).value = int_max_marks

        # Set formulas for student rows
        for idx in range(num_students):
            r = 7 + idx
            sheet.cell(row=r, column=1).value = f"='Students Details'!A{r}"
            sheet.cell(row=r, column=2).value = f"=IF('Students Details'!B{r}=\"\",\"\",'Students Details'!B{r})"
            sheet.cell(row=r, column=3).value = f"=IF('Students Details'!C{r}=\"\",\"\",'Students Details'!C{r})"
            for co in range(1, 7):
                # 'I Component' R is column 18, matching 17 + co
                sheet.cell(row=r, column=3 + co).value = f"=IF('I Component'!{col_letter_of(17+co)}{8+idx}=0,\"\",'I Component'!{col_letter_of(17+co)}{8+idx})"
            sheet.cell(row=r, column=10).value = f"=SUM(D{r}:I{r})"

        # Clear remaining student rows formulas
        for r in range(7 + num_students, 207):
            for c in range(1, 11):
                sheet.cell(row=r, column=c).value = None

    # For External (Grade-based)
    external_paper = papers.get('external', {})
    ext_co_limits = external_paper.get('co_limits', {})
    ext_max_marks = external_paper.get('max_marks', 70)

    if 'E' in wb.sheetnames:
        sheet = wb['E']
        sheet['O4'].value = ext_max_marks

        # Set max marks & percent weights per CO
        for co in range(1, 7):
            co_max = ext_co_limits.get(str(co), 0)
            sheet.cell(row=18 + co, column=16).value = co_max # Col P
            sheet.cell(row=18 + co, column=15).value = (co_max / ext_max_marks * 100) if ext_max_marks else 0 # Col O

        # Set student row details & grade
        for idx, student in enumerate(students):
            r = 7 + idx
            sheet.cell(row=r, column=1).value = f"='Students Details'!A{r}"
            sheet.cell(row=r, column=2).value = f"=IF('Students Details'!B{r}=\"\",\"\",'Students Details'!B{r})"
            sheet.cell(row=r, column=3).value = f"=IF('Students Details'!C{r}=\"\",\"\",'Students Details'!C{r})"

            # Formulas for CO marks & lookup
            sheet.cell(row=r, column=10).value = f"=IF(K{r}=\"\",\"\",LOOKUP(K{r},$M$5:$M$12,$O$5:$O$12))" # Col J
            sheet.cell(row=r, column=4).value = f"=IF(K{r}=\"\",\"\",$O$19%*J{r})" # Col D
            sheet.cell(row=r, column=5).value = f"=IF(K{r}=\"\",\"\",$O$20%*J{r})" # Col E
            sheet.cell(row=r, column=6).value = f"=IF(K{r}=\"\",\"\",$O$21%*J{r})" # Col F
            sheet.cell(row=r, column=7).value = f"=IF($N$22=\"\",\"\",IF(K{r}=\"\",\"\",$O$22%*J{r}))" # Col G
            sheet.cell(row=r, column=8).value = f"=IF($N$23=\"\",\"\",IF(K{r}=\"\",\"\",$O$23%*J{r}))" # Col H
            sheet.cell(row=r, column=9).value = f"=IF($N$24=\"\",\"\",IF(K{r}=\"\",\"\",$O$24%*J{r}))" # Col I

            enroll = student.get('enrollment_no', '')
            grade = student_marks.get(enroll, {}).get('external_grade', '')
            sheet.cell(row=r, column=11).value = grade if grade else None # Col K

        # Clear remaining student rows formulas
        for r in range(7 + num_students, 207):
            for c in range(1, 12):
                sheet.cell(row=r, column=c).value = None

    # For Viva (Grade-based)
    viva_paper = papers.get('viva', {})
    viva_co_limits = viva_paper.get('co_limits', {})
    viva_max_marks = viva_paper.get('max_marks', 30)

    if 'V' in wb.sheetnames:
        sheet = wb['V']
        sheet['O4'].value = f"='CO-PO Mapping'!G17"

        # Set max marks & percent weights per CO
        for co in range(1, 7):
            co_max = viva_co_limits.get(str(co), 0)
            sheet.cell(row=18 + co, column=15).value = co_max # Col O
            sheet.cell(row=18 + co, column=14).value = (co_max / viva_max_marks * 100) if viva_max_marks else 0 # Col N

        # Set student row details & grade
        for idx, student in enumerate(students):
            r = 7 + idx
            sheet.cell(row=r, column=1).value = f"='Students Details'!A{r}"
            sheet.cell(row=r, column=2).value = f"=IF('Students Details'!B{r}=\"\",\"\",'Students Details'!B{r})"
            sheet.cell(row=r, column=3).value = f"=IF('Students Details'!C{r}=\"\",\"\",'Students Details'!C{r})"

            # Formulas for CO marks & lookup
            sheet.cell(row=r, column=10).value = f"=IF(K{r}=\"\",\"\",LOOKUP(K{r},$M$5:$M$12,$O$5:$O$12))" # Col J
            sheet.cell(row=r, column=4).value = f"=IF(K{r}=\"\",\"\",$N$19%*J{r})" # Col D
            sheet.cell(row=r, column=5).value = f"=IF(K{r}=\"\",\"\",$N$20%*J{r})" # Col E
            sheet.cell(row=r, column=6).value = f"=IF(K{r}=\"\",\"\",$N$21%*J{r})" # Col F
            sheet.cell(row=r, column=7).value = f"=IF($N$22=\"\",\"\",IF(K{r}=\"\",\"\",$N$22%*J{r}))" # Col G
            sheet.cell(row=r, column=8).value = f"=IF($N$23=\"\",\"\",IF(K{r}=\"\",\"\",$N$23%*J{r}))" # Col H
            sheet.cell(row=r, column=9).value = f"=IF($N$24=\"\",\"\",IF(K{r}=\"\",\"\",$N$24%*J{r}))" # Col I

            enroll = student.get('enrollment_no', '')
            grade = student_marks.get(enroll, {}).get('viva_grade', '')
            sheet.cell(row=r, column=11).value = grade if grade else None # Col K

        # Clear remaining student rows formulas
        for r in range(7 + num_students, 207):
            for c in range(1, 12):
                sheet.cell(row=r, column=c).value = None

    # 7. Update 'COPO_Th_Pr_Oth' sheet with structured unit headers & lecture subtopics
    if 'COPO_Th_Pr_Oth' in wb.sheetnames:
        sheet = wb['COPO_Th_Pr_Oth']
        # Clear existing planning rows (from row 5 to 111, cols A to I)
        for r in range(5, 112):
            for c in range(1, 10):
                sheet.cell(row=r, column=c).value = None

        lecture_plan = data.get('lecture_plan', [])
        import re

        # Group lectures by unit number
        unit_groups = {}
        for lp in lecture_plan:
            unit_str = str(lp.get('unit', 'Unit 1'))
            match = re.search(r'\d+', unit_str)
            unit_num = int(match.group()) if match else 1
            if unit_num not in unit_groups:
                unit_groups[unit_num] = []
            unit_groups[unit_num].append(lp)

        default_unit_headers = {
            1: 5,
            2: 13,
            3: 20,
            4: 27,
            5: 36,
            6: 46
        }

        current_row = 5
        for unit_num in sorted(unit_groups.keys()):
            items = unit_groups[unit_num]
            if current_row > 110:
                break
            
            target_header_row = max(default_unit_headers.get(unit_num, current_row), current_row)
            
            # Write Unit Header Row
            sheet.cell(row=target_header_row, column=1).value = f"Unit {unit_num}"
            first_topic = items[0].get('topic', '') if items else ''
            sheet.cell(row=target_header_row, column=2).value = f"Unit {unit_num}: {first_topic}"
            
            sub_row = target_header_row + 1
            for sub_idx, lp in enumerate(items, start=1):
                if sub_row > 111:
                    break
                sub_unit_str = f"{unit_num}.{sub_idx}"
                topic_str = str(lp.get('topic', ''))
                co_num = int(lp.get('co_number', 1))

                sheet.cell(row=sub_row, column=1).value = sub_unit_str
                sheet.cell(row=sub_row, column=2).value = topic_str
                sheet.cell(row=sub_row, column=3).value = "Th"
                sheet.cell(row=sub_row, column=4).value = 1
                if 1 <= co_num <= 5:
                    sheet.cell(row=sub_row, column=4 + co_num).value = "1"
                sub_row += 1

            current_row = sub_row

    # Replace [1]Input Data with subject/faculty metadata in Lesson Planning & Lab Planning
    planning_sheets = ['Lesson Planning', 'Lab Planning']
    for p_name in planning_sheets:
        if p_name in wb.sheetnames:
            p_sheet = wb[p_name]
            p_sheet['B3'].value = f"{subject_name} ({subject_code})"
            p_sheet['B4'].value = faculty_name
            p_sheet['B5'].value = f"Sem - {sem_number} ({academic_year} {session})"

    # Save output workbook
    wb.save(output_path)
    print(f"SUCCESS: Generated Excel report at {output_path}")

if __name__ == "__main__":
    main()
